import logging
import os
import time
import openpyxl

start_time = time.perf_counter()

wb = openpyxl.load_workbook("employee_details.xlsx")

sheet = wb["Sheet1"]
header = []
# data = []
ws = wb.active
# print(ws)
rows_count = 0

for row in ws:
    if not all([cell.value is None for cell in row]):
        rows_count += 1

columns = sheet.max_column
print(columns)
print(rows_count)

for cols in range(1, columns + 1):
    header.append(sheet.cell(1, cols).value)

for row in range(2, rows_count + 1):
    data = []
    for cols in range(1, columns + 1):
        data.append(sheet.cell(row, cols).value)

    print(header)
    print(data)
    print("")

    context = dict(zip(header,  data))
    print(context)
    print("===")

# print(header)
# ==========================================================================
#
# import openpyxl
# from docxtpl import DocxTemplate
#
# wb = openpyxl.load_workbook("employee_details.xlsx")
# thisDoc = DocxTemplate("Letter_of_Training.docx")
#
# sheet = wb["Sheet1"]
# ws = wb.active
# columns_count = sheet.max_column
# rows_count = 0
#
# for row in ws:
#     if not all([cell.value is None for cell in row]):
#         rows_count += 1
#
# cell_range = 'A2:G11'
# header = []
#
# for col in range(1, columns_count + 1):
#     header.append(sheet.cell(1, col).value)
#
# data = [[cell.value for cell in row] for row in ws[cell_range]]
# context = {}
#
# for i in data:
#     print(i)
#     print(header)
#     context = dict(zip(header, map(str, i)))
#     print(context)
#     print("==========")
#     thisDoc.render(context)
#     thisDoc.save(f'./letters/{context["employee_name"]}_training_letter.docx')
#
