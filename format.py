import openpyxl
import numpy as np


def format_xl(file_name):
    wb2 = openpyxl.load_workbook(file_name)

    for sheet in wb2.worksheets:
        print('Your currently in ', sheet)
        max_row_in_sheet = sheet.max_row
        max_col_in_sheet = sheet.max_column
        print(max_row_in_sheet, max_col_in_sheet)

        array_3 = np.array([])
        array_4 = np.array([])

        r = 1  # initially declaring row as 1
        c = 1  # initially declaring column as 1
        for r in range(1, max_row_in_sheet + 1):
            array_1 = np.array([])
            array_2 = np.array([])
            for c in range(1, max_col_in_sheet + 1):
                if sheet.cell(row=r, column=c).value == None:
                    array_1 = np.append(array_2, c)
                    array_2 = array_1
            if len(array_1) == max_col_in_sheet:
                array_3 = np.append(array_4, r)
                array_4 = array_3
                array_3 = array_3.astype(int)
        if len(array_3) != 0:
            index_of_last_array_element = len(array_3) - 1
            while index_of_last_array_element != -1:
                sheet.delete_rows(array_3[index_of_last_array_element], 1)
                index_of_last_array_element = index_of_last_array_element \
                                              - 1

        max_row_in_sheet = sheet.max_row  # maximum entered row
        max_col_in_sheet = sheet.max_column  # maximum entered column

        print('Maximum Rows and Cols after Removing')
        print(max_row_in_sheet, max_col_in_sheet)
        print('======================================')
        col_arr = []
        for x in range(1, sheet.max_column + 1):
            col_arr.append(0)

        for r in range(1, max_row_in_sheet + 1):
            array_1 = np.array([])
            array_2 = np.array([])
            for c in range(1, max_col_in_sheet + 1):
                if sheet.cell(row=r, column=c).value == None:
                    array_1 = np.append(array_2, c)
                    array_2 = array_1
                    col_arr[c - 1] += 1
        print(col_arr)

        array_2 = [int(x) for x in array_2]
        print(len(array_2))
        print(array_2)
        if len(array_2) != 0:
            index = len(array_2) - 1
            print(index)
            while index != -1:
                temp = array_2[index]

                # print(temp)

                sheet.delete_cols(temp, 1)
                index = index - 1

    wb2.save('employee_details.xlsx')
