import time
import openpyxl
import os.path
from docxtpl import DocxTemplate
from docx2pdf import convert
import format
import logging


def generate():
    start_time = time.perf_counter()
    logging.basicConfig(filename='app.log', filemode='a', format='%(name)s - %(levelname)s - %(message)s')
    wb = openpyxl.load_workbook("employee_details.xlsx")
    thisDoc = DocxTemplate("Letter_of_Training.docx")

    sheet = wb["Sheet1"]
    header = []
    ws = wb.active
    # print(ws)
    rows_count = sheet.max_row

    # for row in ws:
    #     if not all([cell.value is None for cell in row]):
    #         rows_count += 1

    # check for empty rows in between

    columns = sheet.max_column
    print(columns)
    print(rows_count)

    # {emp_name:"souvik", "emp_salary":2000,}

    for cols in range(1, columns + 1):
        header.append(sheet.cell(1, cols).value)

    for row in range(2, rows_count + 1):
        data = []
        for cols in range(1, columns + 1):
            data.append(sheet.cell(row, cols).value)

        print(data)
        print(header)
        print("")

        context = dict(zip(header, data))
        print(context)
        print("===")
        try:
            if os.path.isfile(f'./letters/{context["employee_name"]}_training_letter.docx') \
                    or os.path.isfile(f'./letters/{context["employee_name"]}_training_letter.pdf'):
                logging.warning("File already exists")
            else:
                try:
                    thisDoc.render(context)
                    thisDoc.save(f'./letters/{context["employee_name"]}_training_letter.docx')
                    logging.warning(f"new file created for => {context['employee_name']}")
                    try:
                        # convert to pdf
                        input_file = f'./letters/{context["employee_name"]}_training_letter.docx'
                        output_file = f'./letters/{context["employee_name"]}_training_letter.pdf'
                        convert(input_file, output_file)
                        logging.warning(f".PDF file generated for {context['employee_name']}")
                        # convert("./letters")
                    except Exception as e:
                        logging.error(f"problem in pdf generation: {e}")
                except:
                    logging.error("problem in .docx generation")
        except Exception as e:
            logging.error(e)

        # delete .docx files and only retain those pdf files
        # try:
        #     os.remove(f'./letters/{context["employee_name"]}_training_letter.docx')
        # except OSError as e:
        #     logging.error(e)

    end_time = time.perf_counter()
    print(end_time - start_time, "seconds")


def main():
    format.format_xl("employee_details.xlsx")  # format the Excel file
    generate()


if __name__ == "__main__":
    main()
